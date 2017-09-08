# -*- coding: utf-8 -*-
#
# Copyright (C) 2012 Daniel Rolls <drolls@maxeler.com>
# Copyright (C) 2012 Maxeler Technologies Inc
# Derived from Christopher Lenz's ExcelViewerPlugin
#   Copyright (C) 2006 Christopher Lenz <cmlenz@gmx.de>
#   All rights reserved.
#
# This software is licensed as described in the file COPYING, which
# you should have received as part of this distribution.

import os
import tempfile

from trac.core import Component, implements
from trac.mimeview.api import IHTMLPreviewRenderer
from trac.util import escape

__all__ = ['ExcelXSLTRenderer']

try:
    import openpyxl
except ImportError:
    have_openpyxl = False
else:
    have_openpyxl = True


class ExcelXSLTRenderer(Component):
    implements(IHTMLPreviewRenderer)

    MIME_TYPES = ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                  'application/vnd.ms-excel.sheet.macroEnabled.12')

    # IHTMLPreviewRenderer methods

    def get_quality_ratio(self, mimetype):
        if have_openpyxl and mimetype in self.MIME_TYPES:
            return 2
        return 0

    def render(self, req, mimetype, content, filename=None, url=None):
        if hasattr(content, 'read'):
            content = content.read()
        fd, path = tempfile.mkstemp(suffix='.xlsx', text=True)
        os.write(fd, content)
        os.close(fd)
        book = openpyxl.load_workbook(path, read_only=False, data_only=True)
        buf = []
        for sheet_name in book.get_sheet_names():
            sheet = book.get_sheet_by_name(sheet_name)
            if sheet.max_row == 1 and sheet.max_column == 1 and \
                    not sheet['A1'].value:
                continue  # Skip empty sheet
            buf.append(u'<table class="listing"><caption>%s</caption>\n'
                       % escape(sheet_name))
            buf.append(u'<tbody>')
            for row_idx, row in enumerate(sheet.rows):
                buf.append(u'<tr class="%s">'
                           % ('odd' if row_idx % 2 else 'even'))
                for cell in row:
                    val = cell.value if cell.value is not None else u''
                    buf.append(u'<td>%s</td>' % val)
                buf.append(u'</tr>\n')
            buf.append(u'</tbody>')
            buf.append(u'</table>\n')
        os.unlink(path)
        return u''.join(buf) if buf else "Empty workbook"
